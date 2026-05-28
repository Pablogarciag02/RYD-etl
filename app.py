"""RYD — Streamlit app for daily per-table Excel uploads.

User flow:
    1. Pick which table is in the file (radio)
    2. Upload an .xlsx
    3. App validates → loads raw → refreshes dims → refreshes that table's facts
    4. Success message → "Upload another" resets the form

Run:
    source .venv/bin/activate
    streamlit run app.py

This file is the UI shell. The actual extract/load/transform/validate/audit logic
lives in the `etl/` package and is shared with the CLI pipeline.
"""

import hmac
import os
import sys
import tempfile
from datetime import datetime

import psycopg2
import streamlit as st
from dotenv import load_dotenv
from openpyxl import load_workbook

# Make the etl package importable when running `streamlit run app.py`.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from etl.extract import (
    SHEET_MAP,
    extract_finance_applications,
    extract_market_prices,
    extract_sales,
    extract_claims,
    extract_inegi_sales,
)
from etl.load_raw import (
    bulk_insert,
    count_rows_for_batch,
    create_import_batch,
    fail_import_batch,
    find_completed_batch,
    get_batch_counts,
    update_batch_status,
)
from etl.validate import validate_sheet_contract
from etl.audit import (
    create_sheet_run,
    record_import_error,
    record_validation_result,
    update_sheet_run_status,
)
from etl.transform import (
    populate_dimensions,
    populate_fact_finance_applications,
    populate_fact_market_prices,
    populate_fact_sales,
    populate_fact_claims,
    populate_fact_market_sales_inegi,
)

load_dotenv()


# ══════════════════════════════════════════════════════════════
# Secret access — Streamlit Cloud secrets, falling back to env
# ══════════════════════════════════════════════════════════════

def _secret(key, default=None):
    """Prefer st.secrets (Streamlit Cloud). Fall back to env (local .env)."""
    try:
        if key in st.secrets:
            return st.secrets[key]
    except (FileNotFoundError, Exception):
        pass
    return os.getenv(key, default)


def get_connection():
    conn = psycopg2.connect(
        host=_secret("DB_HOST"),
        port=int(_secret("DB_PORT", 5432)),
        dbname=_secret("DB_NAME"),
        user=_secret("DB_USER"),
        password=_secret("DB_PASSWORD"),
        sslmode="require",
    )
    # Safety brake: any single statement that runs longer than 60s
    # aborts cleanly instead of dragging the whole instance into
    # IO/CPU exhaustion. Set via SET (not connect options=) so it
    # works through Supabase's Supavisor pooler in any mode.
    # The commit() is required: SET opens an implicit transaction in
    # psycopg2's default non-autocommit mode, and returning a connection
    # with an open transaction breaks any subsequent `conn.autocommit = ...`
    # assignment (psycopg2 calls set_session() which forbids in-transaction).
    with conn.cursor() as c:
        c.execute("SET statement_timeout = 60000")
    conn.commit()
    return conn


# ══════════════════════════════════════════════════════════════
# Table registry — one entry per user-selectable upload target.
# `template` is the canonical sheet name that validate.SHEET_CONTRACTS
# is keyed on; `extractor` and `target_table` come from etl.extract.SHEET_MAP;
# `fact_fn` is the per-table fact loader from etl.transform.
# ══════════════════════════════════════════════════════════════

TABLES = {
    "finance": {
        "label": "Finance Applications (DBSolicitudes_Cetelem)",
        "template": "DBSolicitudes_Cetelem",
        "fact_fn": populate_fact_finance_applications,
    },
    "prices": {
        "label": "Market Prices (DBPreciosMexico_ConMG)",
        "template": "DBPreciosMexico_ConMG",
        "fact_fn": populate_fact_market_prices,
    },
    "sales": {
        "label": "Sales (DBVentas_ConMG)",
        "template": "DBVentas_ConMG",
        "fact_fn": populate_fact_sales,
    },
    "claims": {
        "label": "Claims (DBSiniestros_Marsh)",
        "template": "DBSiniestros_Marsh",
        "fact_fn": populate_fact_claims,
    },
    "inegi": {
        "label": "INEGI Sales (BaseINEGIAutosLigerosMexico)",
        "template": "BaseINEGIAutosLigerosMexico",
        "fact_fn": populate_fact_market_sales_inegi,
    },
}


# ══════════════════════════════════════════════════════════════
# Upload helpers
# ══════════════════════════════════════════════════════════════

def _save_uploaded_bytes_as_named_sheet(file_bytes, sheet_name):
    """
    Persist the uploaded bytes to a temp .xlsx and ensure a sheet named
    `sheet_name` exists. validate.validate_sheet_contract and the
    extractors look up sheets by name; if the user uploaded a workbook
    whose first sheet is named e.g. "Hoja1", we rename it to match the
    contract before processing. Returns the temp file path.
    """
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    with open(path, "wb") as f:
        f.write(file_bytes)

    # Cheap streaming check first — load_workbook() in default (writable)
    # mode pulls the entire workbook into RAM, which can OOM a small
    # Streamlit Cloud container on large files (e.g. a 42 MB workbook can
    # balloon past 1 GB in memory). Only fall back to writable mode if a
    # rename is actually required; for the common case where the uploaded
    # workbook's first sheet already matches the contract name, we stay
    # in streaming mode and never load the full workbook.
    wb_ro = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_already_named = sheet_name in wb_ro.sheetnames
    finally:
        wb_ro.close()

    if not sheet_already_named:
        wb = load_workbook(path)
        try:
            wb.worksheets[0].title = sheet_name
            wb.save(path)
        finally:
            wb.close()

    return path


def _record_validation_checks(cur, batch_id, sheet_run_id, sheet_name, validation):
    for check in validation.get("checks", []):
        record_validation_result(
            cur=cur,
            import_batch_id=batch_id,
            sheet_run_id=sheet_run_id,
            source_sheet_name=sheet_name,
            validation_stage=check["validation_stage"],
            validation_name=check["validation_name"],
            status=check["status"],
            severity=check.get("severity", "error"),
            expected_value=str(check["expected_value"]) if check.get("expected_value") is not None else None,
            actual_value=str(check["actual_value"]) if check.get("actual_value") is not None else None,
            details=check.get("details", {}),
        )


def run_upload(file_obj, filename, table_code):
    """Process one uploaded file for one table. Returns a stats dict.

    Structure (per the 2026-05-28 refactor): the upload is split into
    PHASES with commits between, so a long upload no longer requires one
    multi-hour transaction.

      PHASE 1 — idempotency check (read-only, no writes)
      PHASE 2 — batch + sheet_run creation + validation (small writes,
                each phase committed independently)
      PHASE 3 — raw load via bulk_insert which commits every batch_size
                rows in autocommit mode; on return raw is durable
      PHASE 4 — transforms (dims + this table's fact) in a fresh
                transaction. If they fail, the raw rows for this batch
                are deleted in a follow-up tx so retries start clean.

    Previously the whole upload was wrapped in a single transaction,
    which on Streamlit Cloud + Supabase would silently roll back the
    entire upload if the connection dropped mid-flight (~2h timeouts
    we don't control). With chunked commits, a drop loses at most the
    in-flight chunk.
    """
    cfg = TABLES[table_code]
    sheet_name = cfg["template"]
    extractor, target_table = SHEET_MAP[sheet_name]
    fact_fn = cfg["fact_fn"]

    file_bytes = file_obj.read()
    tmp_path = _save_uploaded_bytes_as_named_sheet(file_bytes, sheet_name)

    conn = get_connection()
    conn.autocommit = False
    started_at = datetime.utcnow()

    batch_id = None
    sheet_run_id = None

    try:
        cur = conn.cursor()

        # ── PHASE 1: idempotency check ──
        existing_batch_id = find_completed_batch(cur, tmp_path, sheet_name)
        if existing_batch_id is not None:
            counts = get_batch_counts(cur, existing_batch_id) or {}
            db_count = count_rows_for_batch(cur, target_table, existing_batch_id)
            sheet_run_id = create_sheet_run(
                cur=cur,
                import_batch_id=existing_batch_id,
                source_file_name=filename,
                source_sheet_name=sheet_name,
                target_table=target_table,
            )
            update_sheet_run_status(
                cur=cur,
                sheet_run_id=sheet_run_id,
                status="skipped",
                rows_detected=counts.get("rows_detected", 0),
                rows_loaded=counts.get("rows_loaded", 0),
                rows_failed=counts.get("rows_failed", 0),
                notes="Skipped: same file checksum + sheet was already processed.",
                started_at=started_at,
                finished_at=datetime.utcnow(),
                metadata={
                    "existing_batch_id": str(existing_batch_id),
                    "reconciliation": {
                        "loaded": counts.get("rows_loaded", 0),
                        "db_count_for_batch": db_count,
                        "ok": db_count == counts.get("rows_loaded", 0),
                    },
                },
            )
            conn.commit()
            return {
                "skipped": True,
                "batch_id": str(existing_batch_id),
                "raw_loaded": counts.get("rows_loaded", 0),
                "raw_failed": counts.get("rows_failed", 0),
                "facts_inserted": 0,
            }

        # ── PHASE 2: batch + sheet_run creation + validation ──
        batch_id = create_import_batch(cur, tmp_path, sheet_name)
        sheet_run_id = create_sheet_run(
            cur=cur,
            import_batch_id=batch_id,
            source_file_name=filename,
            source_sheet_name=sheet_name,
            target_table=target_table,
        )
        conn.commit()

        update_sheet_run_status(cur=cur, sheet_run_id=sheet_run_id, status="validating")
        conn.commit()

        validation = validate_sheet_contract(tmp_path, sheet_name)
        _record_validation_checks(cur, batch_id, sheet_run_id, sheet_name, validation)
        conn.commit()  # commit validation results so they survive later drops

        if not validation["ok"]:
            record_import_error(
                cur=cur,
                import_batch_id=batch_id,
                sheet_run_id=sheet_run_id,
                source_file_name=filename,
                source_sheet_name=sheet_name,
                target_table=target_table,
                error_stage="pre_ingestion_validation",
                error_type=validation["error_type"],
                error_message=validation["message"],
                error_details=validation.get("details", {}),
            )
            fail_import_batch(cur, batch_id, validation["message"])
            update_sheet_run_status(
                cur=cur,
                sheet_run_id=sheet_run_id,
                status="failed",
                rows_detected=0,
                rows_loaded=0,
                rows_failed=0,
                notes=validation["message"],
                started_at=started_at,
                finished_at=datetime.utcnow(),
                metadata=validation.get("details", {}),
            )
            conn.commit()
            raise ValueError(validation["message"])

        update_sheet_run_status(cur=cur, sheet_run_id=sheet_run_id, status="processing")
        conn.commit()

        # ── PHASE 3: raw load with chunked commits ──
        # bulk_insert toggles conn.autocommit internally so each chunk is
        # its own committed transaction; on return raw is already durable
        # on disk regardless of what happens next.
        rows = extractor(tmp_path)
        detected, loaded, failed = bulk_insert(cur, conn, target_table, rows, batch_id)

        # ── PHASE 4: transforms (fresh transaction) ──
        # If anything in this block fails, raw is still safe; the except
        # below cleans up the orphaned raw rows for this batch and marks
        # the batch failed so retries can start clean.
        try:
            db_count = count_rows_for_batch(cur, target_table, batch_id)
            recon_ok = db_count == loaded

            populate_dimensions(cur)
            # Pass batch_id so the fact load only processes THIS batch's
            # raw rows instead of rescanning the whole raw table.
            fact_fn(cur, batch_id=batch_id)

            update_batch_status(cur, batch_id, detected, loaded, failed)
            update_sheet_run_status(
                cur=cur,
                sheet_run_id=sheet_run_id,
                status="success" if failed == 0 else "partial_success",
                rows_detected=detected,
                rows_loaded=loaded,
                rows_failed=failed,
                notes=None if recon_ok else f"Row count mismatch: loaded={loaded} db_count={db_count}",
                started_at=started_at,
                finished_at=datetime.utcnow(),
                metadata={
                    "reconciliation": {
                        "detected": detected,
                        "loaded": loaded,
                        "failed": failed,
                        "db_count_for_batch": db_count,
                        "ok": recon_ok,
                    },
                },
            )
            conn.commit()

            return {
                "skipped": False,
                "batch_id": str(batch_id),
                "raw_loaded": loaded,
                "raw_failed": failed,
                "facts_inserted": cur.rowcount,
            }

        except Exception as transform_err:
            # Roll back any partial transform work, then clean up the
            # raw rows that were committed in Phase 3 (otherwise they
            # sit orphaned in raw forever) and mark the batch failed.
            conn.rollback()
            try:
                cur.execute(
                    f"DELETE FROM {target_table} WHERE import_batch_id = %s",
                    (str(batch_id),),
                )
                fail_import_batch(cur, batch_id, f"Transform failed: {transform_err}")
                update_sheet_run_status(
                    cur=cur,
                    sheet_run_id=sheet_run_id,
                    status="failed",
                    rows_detected=detected,
                    rows_loaded=0,
                    rows_failed=loaded,
                    notes=(
                        f"Transform stage failed: {transform_err}. "
                        f"Raw rows for this batch were deleted to keep state consistent."
                    ),
                    started_at=started_at,
                    finished_at=datetime.utcnow(),
                    metadata={"phase": "transform", "exception_class": type(transform_err).__name__},
                )
                conn.commit()
            except Exception:
                conn.rollback()
            raise

    except Exception:
        # Anything in Phase 1/2 — uncommitted state is rolled back; Phase
        # 3/4 have their own handling above. Re-raise so the UI shows it.
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        conn.close()
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


# ══════════════════════════════════════════════════════════════
# Password gate
# ══════════════════════════════════════════════════════════════

def check_password():
    """Block the app behind a shared password. Call before rendering UI."""
    expected = _secret("APP_PASSWORD")
    if not expected:
        st.error("APP_PASSWORD is not configured. Set it in Streamlit secrets "
                 "(or a local .env file) before running the app.")
        st.stop()

    def _on_submit():
        entered = st.session_state.get("pw_input", "") or ""
        # Compare as bytes so non-ASCII (e.g. smart-quote autocorrect) returns
        # False instead of raising TypeError.
        if hmac.compare_digest(entered.encode("utf-8"), str(expected).encode("utf-8")):
            st.session_state["auth_ok"] = True
            del st.session_state["pw_input"]
        else:
            st.session_state["auth_ok"] = False

    if st.session_state.get("auth_ok"):
        return

    st.text_input("Password", type="password", on_change=_on_submit, key="pw_input")
    if "auth_ok" in st.session_state and not st.session_state["auth_ok"]:
        st.error("Incorrect password.")
    st.stop()


# ══════════════════════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════════════════════

st.set_page_config(page_title="RYD — Data Upload", layout="centered")
check_password()
st.title("RYD — Data Upload")
st.caption("Upload one Excel file per table, one at a time.")

if "history" not in st.session_state:
    st.session_state.history = []
if "form_nonce" not in st.session_state:
    st.session_state.form_nonce = 0

label_to_code = {cfg["label"]: code for code, cfg in TABLES.items()}

with st.form(key=f"upload_form_{st.session_state.form_nonce}", clear_on_submit=True):
    table_label = st.radio(
        "Which table is in this file?",
        options=list(label_to_code.keys()),
        index=0,
    )
    uploaded = st.file_uploader("Excel file (.xlsx)", type=["xlsx"])
    submit = st.form_submit_button("Process upload", type="primary")

if submit:
    if uploaded is None:
        st.error("Please choose a file before submitting.")
    else:
        table_code = label_to_code[table_label]
        with st.spinner(f"Loading {table_label}…"):
            try:
                stats = run_upload(uploaded, uploaded.name, table_code)
                st.session_state.history.append({"label": table_label, **stats})
                if stats.get("skipped"):
                    st.info(
                        f"**Already loaded.** `{table_label}` matches an earlier successful upload "
                        f"(same file checksum + sheet). Nothing to do.\n\n"
                        f"- Existing batch ID: `{stats['batch_id']}`"
                    )
                else:
                    st.success(
                        f"**Success!** `{table_label}` uploaded.\n\n"
                        f"- Raw rows loaded: **{stats['raw_loaded']:,}**\n"
                        f"- Raw rows failed: **{stats['raw_failed']:,}**\n"
                        f"- Batch ID: `{stats['batch_id']}`"
                    )
            except Exception as e:
                st.error(f"Upload failed: {e}")
                # Show the full Python traceback in an expander so we can
                # diagnose errors that the Streamlit Cloud log viewer truncates.
                with st.expander("Technical traceback (click to expand)"):
                    import traceback
                    st.code(traceback.format_exc())

if st.button("Upload another file"):
    st.session_state.form_nonce += 1
    st.rerun()

if st.session_state.history:
    with st.expander(f"This session's uploads ({len(st.session_state.history)})", expanded=False):
        for u in reversed(st.session_state.history):
            tag = "↩ skipped" if u.get("skipped") else "✓ loaded"
            st.text(
                f"{tag}  {u['label']}  —  "
                f"{u.get('raw_loaded', 0):,} raw  "
                f"(batch {u['batch_id'][:8]})"
            )
