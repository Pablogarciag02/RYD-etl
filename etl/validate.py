# etl/validate.py
"""
Validation layer for recurring workbook ingestion.

Checks:
- workbook can be opened
- required sheet exists
- sheet header row exists
- expected headers exist
- expected header order matches exactly
- key column exists
- key column non-empty ratio is acceptable
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional
from openpyxl import load_workbook


def normalize_header(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "Á": "a",
        "É": "e",
        "Í": "i",
        "Ó": "o",
        "Ú": "u",
        "ñ": "n",
        "Ñ": "n",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    return " ".join(text.lower().split())


# Canonical recurring ingestion contracts.
# CatalogoTiendas ConMG and Resumen are intentionally excluded (master/manual data).
# BaseLeads1 and BaseLeads2 are also excluded — leads ingestion no longer flows through this pipeline.
SHEET_CONTRACTS: Dict[str, Dict[str, Any]] = {
    "DBSolicitudes_Cetelem": {
        "target_table": "raw.raw_finance_applications",
        "expected_headers_in_order": [
            "Fecha OK",
            "Cuenta Aprobado",
            "Cuenta Base",
            "Cuenta ANF",
            "Distribuidor_OK",
            "Folio",
            "Mes Solicitud",
            "Fecha y Hora Recibida",
            "Tipo_Auto",
            "Status",
            "Substatus",
            "Modelo",
            "Carline",
            "Versión",
            "Importe Unidad",
            "Enganche",
            "Porcentaje de Enganche",
            "Monto a Financiar",
            "Codigo Distribuidor",
            "Nombre Distribuidor",
            "Grupo",
            "Marca",
            "Tipo de Persona",
            "Plazo",
            "Producto",
        ],
        "key_header": "Folio",
        "min_key_non_empty_ratio": 0.98,
    },
    "DBPreciosMexico_ConMG": {
        "target_table": "raw.raw_market_prices",
        "expected_headers_in_order": [
            "Mes",
            "Segmento RYD",
            "Marca",
            "Modelo",
            "Version",
            "Body type",
            "Puertas",
            "Model year",
            "Retail price",
            "Cash/Net price",
            "Finance price",
            "Price currency",
        ],
        "key_header": "Modelo",
        "min_key_non_empty_ratio": 0.95,
    },
    "DBVentas_ConMG": {
        "target_table": "raw.raw_sales",
        "expected_headers_in_order": [
            "Month",
            "Group",
            "Dealer",
            "Tipo de Venta",
            "Unidades",
            "Carline",
        ],
        "key_header": "Dealer",
        "min_key_non_empty_ratio": 0.95,
    },
    "DBSiniestros_Marsh": {
        "target_table": "raw.raw_claims",
        "expected_headers_in_order": [
            "Fecha",
            "Póliza",
            "Aseguradora",
            "Programa",
            "Nombre Agencia",
            "Razón Social",
            "Nombre de Agencia 2",
            "Grupo",
            "Cliente",
            "Siniestro",
            "Cobertura",
            "Conductor",
            "Teléfono",
            "Vehículo",
            "Serie",
            "Modelo",
            "Ciudad",
            "Estado",
        ],
        "key_header": "Siniestro",
        "min_key_non_empty_ratio": 0.95,
    },
    "BaseINEGIAutosLigerosMexico": {
        "target_table": "raw.raw_inegi_sales",
        "expected_headers_in_order": [
            "Tema",
            "Año",
            "Mes",
            "Marca",
            "Modelo",
            "Tipo",
            "Segmento",
            "Origen",
            "País origen",
            "Cantidad",
        ],
        "key_header": "Marca",
        "min_key_non_empty_ratio": 0.95,
    },
}


def get_workbook_sheetnames(filepath: str) -> List[str]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def get_sheet_headers(filepath: str, sheet_name: str) -> Optional[List[Optional[str]]]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if first_row is None:
            return []
        return [str(h).strip() if h is not None else None for h in first_row]
    finally:
        wb.close()


def compute_key_non_empty_ratio(filepath: str, sheet_name: str, key_header: str) -> Dict[str, Any]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if first_row is None:
            return {
                "ok": False,
                "total_data_rows": 0,
                "non_empty_key_rows": 0,
                "ratio": 0.0,
                "message": "Sheet has no header row.",
            }

        normalized_headers = [normalize_header(h) for h in first_row]
        normalized_key = normalize_header(key_header)

        if normalized_key not in normalized_headers:
            return {
                "ok": False,
                "total_data_rows": 0,
                "non_empty_key_rows": 0,
                "ratio": 0.0,
                "message": f"Key header '{key_header}' not found.",
            }

        key_idx = normalized_headers.index(normalized_key)

        total_data_rows = 0
        non_empty_key_rows = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # skip fully empty rows
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            total_data_rows += 1

            if key_idx < len(row):
                val = row[key_idx]
                if val is not None and str(val).strip() != "":
                    non_empty_key_rows += 1

        ratio = non_empty_key_rows / total_data_rows if total_data_rows > 0 else 0.0

        return {
            "ok": True,
            "total_data_rows": total_data_rows,
            "non_empty_key_rows": non_empty_key_rows,
            "ratio": ratio,
            "message": "Key-column completeness computed successfully.",
        }
    finally:
        wb.close()


def validate_workbook_open(filepath: str) -> Dict[str, Any]:
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        wb.close()
        return {
            "ok": True,
            "validation_name": "workbook_can_be_opened",
            "message": "Workbook opened successfully.",
            "details": {},
        }
    except Exception as e:
        return {
            "ok": False,
            "validation_name": "workbook_can_be_opened",
            "message": f"Workbook could not be opened: {e}",
            "details": {"exception": str(e)},
        }


def validate_sheet_contract(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """
    Validate a single sheet against the canonical contract.

    Returns:
      {
        "ok": bool,
        "error_type": str | None,
        "message": str,
        "checks": [ ... ],
        "details": { ... }
      }
    """
    checks: List[Dict[str, Any]] = []

    # 1) Workbook can be opened
    workbook_check = validate_workbook_open(filepath)
    checks.append({
        "validation_stage": "pre_ingestion_validation",
        "validation_name": workbook_check["validation_name"],
        "status": "passed" if workbook_check["ok"] else "failed",
        "severity": "error",
        "expected_value": "workbook is readable",
        "actual_value": workbook_check["message"],
        "details": workbook_check["details"],
    })
    if not workbook_check["ok"]:
        return {
            "ok": False,
            "error_type": "invalid_workbook",
            "message": workbook_check["message"],
            "checks": checks,
            "details": workbook_check["details"],
        }

    # 2) Known contract exists
    if sheet_name not in SHEET_CONTRACTS:
        checks.append({
            "validation_stage": "pre_ingestion_validation",
            "validation_name": "known_sheet_contract",
            "status": "failed",
            "severity": "error",
            "expected_value": "sheet in SHEET_CONTRACTS",
            "actual_value": sheet_name,
            "details": {"known_sheets": list(SHEET_CONTRACTS.keys())},
        })
        return {
            "ok": False,
            "error_type": "unknown_sheet_contract",
            "message": f"No validation contract defined for sheet '{sheet_name}'.",
            "checks": checks,
            "details": {"known_sheets": list(SHEET_CONTRACTS.keys())},
        }

    contract = SHEET_CONTRACTS[sheet_name]
    expected_headers = contract["expected_headers_in_order"]

    # 3) Required sheet exists
    workbook_sheets = get_workbook_sheetnames(filepath)
    sheet_exists = sheet_name in workbook_sheets
    checks.append({
        "validation_stage": "pre_ingestion_validation",
        "validation_name": "required_sheet_exists",
        "status": "passed" if sheet_exists else "failed",
        "severity": "error",
        "expected_value": sheet_name,
        "actual_value": sheet_name if sheet_exists else None,
        "details": {"workbook_sheets": workbook_sheets},
    })
    if not sheet_exists:
        return {
            "ok": False,
            "error_type": "missing_sheet",
            "message": f"Required sheet '{sheet_name}' was not found in workbook.",
            "checks": checks,
            "details": {"workbook_sheets": workbook_sheets},
        }

    # 4) Header row exists
    actual_headers = get_sheet_headers(filepath, sheet_name)
    header_row_exists = actual_headers is not None and len(actual_headers) > 0
    checks.append({
        "validation_stage": "pre_ingestion_validation",
        "validation_name": "sheet_header_row_exists",
        "status": "passed" if header_row_exists else "failed",
        "severity": "error",
        "expected_value": "non-empty first row",
        "actual_value": actual_headers,
        "details": {},
    })
    if not header_row_exists:
        return {
            "ok": False,
            "error_type": "missing_header_row",
            "message": f"Sheet '{sheet_name}' does not have a readable header row.",
            "checks": checks,
            "details": {},
        }

    assert actual_headers is not None
    normalized_actual_headers = [normalize_header(h) for h in actual_headers]
    normalized_expected_headers = [normalize_header(h) for h in expected_headers]

    # 5) Expected headers exist
    missing_headers = [
        expected_headers[i]
        for i, h in enumerate(normalized_expected_headers)
        if h not in normalized_actual_headers
    ]
    expected_headers_exist = len(missing_headers) == 0

    checks.append({
        "validation_stage": "pre_ingestion_validation",
        "validation_name": "expected_headers_exist",
        "status": "passed" if expected_headers_exist else "failed",
        "severity": "error",
        "expected_value": expected_headers,
        "actual_value": actual_headers,
        "details": {"missing_headers": missing_headers},
    })
    if not expected_headers_exist:
        return {
            "ok": False,
            "error_type": "missing_required_headers",
            "message": (
                f"Sheet '{sheet_name}' is missing required headers: {missing_headers}"
            ),
            "checks": checks,
            "details": {
                "expected_headers": expected_headers,
                "actual_headers": actual_headers,
                "missing_headers": missing_headers,
            },
        }

    # 6) Expected header order matches exactly
    exact_order_match = normalized_actual_headers == normalized_expected_headers
    checks.append({
        "validation_stage": "pre_ingestion_validation",
        "validation_name": "expected_header_order_matches",
        "status": "passed" if exact_order_match else "failed",
        "severity": "error",
        "expected_value": expected_headers,
        "actual_value": actual_headers,
        "details": {},
    })
    if not exact_order_match:
        return {
            "ok": False,
            "error_type": "header_order_mismatch",
            "message": (
                f"Sheet '{sheet_name}' has unexpected header order. "
                "This file cannot be safely processed with positional extractors."
            ),
            "checks": checks,
            "details": {
                "expected_headers": expected_headers,
                "actual_headers": actual_headers,
            },
        }

    # 7) Key column exists
    key_header = contract["key_header"]
    key_exists = normalize_header(key_header) in normalized_actual_headers
    checks.append({
        "validation_stage": "pre_ingestion_validation",
        "validation_name": "key_column_exists",
        "status": "passed" if key_exists else "failed",
        "severity": "error",
        "expected_value": key_header,
        "actual_value": actual_headers,
        "details": {},
    })
    if not key_exists:
        return {
            "ok": False,
            "error_type": "missing_key_header",
            "message": f"Key header '{key_header}' is missing in sheet '{sheet_name}'.",
            "checks": checks,
            "details": {
                "expected_key_header": key_header,
                "actual_headers": actual_headers,
            },
        }

    # 8) Key column non-empty ratio acceptable
    key_completeness = compute_key_non_empty_ratio(filepath, sheet_name, key_header)
    min_ratio = contract["min_key_non_empty_ratio"]
    ratio_ok = key_completeness["ok"] and key_completeness["ratio"] >= min_ratio

    checks.append({
        "validation_stage": "pre_ingestion_validation",
        "validation_name": "key_column_non_empty_ratio",
        "status": "passed" if ratio_ok else "failed",
        "severity": "error",
        "expected_value": f">= {min_ratio:.0%}",
        "actual_value": f"{key_completeness['ratio']:.2%}",
        "details": key_completeness,
    })
    if not ratio_ok:
        return {
            "ok": False,
            "error_type": "insufficient_key_completeness",
            "message": (
                f"Key header '{key_header}' in sheet '{sheet_name}' has "
                f"non-empty ratio {key_completeness['ratio']:.2%}, below required {min_ratio:.0%}."
            ),
            "checks": checks,
            "details": {
                "key_header": key_header,
                "min_required_ratio": min_ratio,
                **key_completeness,
            },
        }

    return {
        "ok": True,
        "error_type": None,
        "message": f"Sheet '{sheet_name}' passed validation.",
        "checks": checks,
        "details": {
            "sheet_name": sheet_name,
            "target_table": contract["target_table"],
            "key_header": key_header,
            "key_non_empty_ratio": key_completeness["ratio"],
            "total_data_rows": key_completeness["total_data_rows"],
        },
    }


def validate_requested_sheets(filepath: str, sheet_names: List[str]) -> Dict[str, Any]:
    """
    Validate multiple requested sheets.
    Returns a summary with per-sheet results.
    """
    results = {}
    overall_ok = True

    for sheet_name in sheet_names:
        result = validate_sheet_contract(filepath, sheet_name)
        results[sheet_name] = result
        if not result["ok"]:
            overall_ok = False

    return {
        "ok": overall_ok,
        "results": results,
    }


def get_expected_ingestion_sheets() -> List[str]:
    return list(SHEET_CONTRACTS.keys())