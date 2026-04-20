"""Extract layer: reads Excel sheets and yields dicts mapped to raw table columns."""

import json
from datetime import datetime, timedelta
from openpyxl import load_workbook


def _str(val):
    if val is None:
        return None
    return str(val).strip() or None


def _excel_serial_to_datetime(serial):
    """Convert Excel serial date number to datetime."""
    if serial is None:
        return None
    try:
        serial = float(serial)
        base = datetime(1899, 12, 30)
        return base + timedelta(days=serial)
    except (ValueError, TypeError):
        return None


def _dt_str(val):
    """Convert a datetime (or Excel serial) to ISO string."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        val = _excel_serial_to_datetime(val)
    if isinstance(val, datetime):
        return val.isoformat()
    return str(val).strip() or None


def _num_str(val):
    """Store a number as text for raw tables."""
    if val is None:
        return None
    return str(val)


def read_sheets(filepath):
    """Return a dict of {sheet_name: list_of_header_names} for inspection."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
        if first_row:
            result[name] = [str(h) if h else None for h in first_row]
    wb.close()
    return result


def extract_leads(filepath):
    """BaseLeads1 → raw.raw_leads rows."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["BaseLeads1"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_num, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        yield {
            "source_row_number": row_num,
            "aglead": _str(row[0]),
            "distribuidor": _str(row[1]),
            "fecha_origen_raw": _dt_str(row[2]),
            "hora_origen_raw": _num_str(row[3]),
            "grupo": _str(row[17]),
            "marca": _str(row[16]),
            "producto": _str(row[5]),
            "asesor": _str(row[7]),
            "campania": _str(row[10]),
            "subcampania": _str(row[11]),
            "fuente": _str(row[8]),
            "medio_atencion": _str(row[14]),
            "titulo_lead": _str(row[15]),
            "temperatura": _str(row[6]),
            "estatus": _str(row[9]),
            "motivo_finalizacion": _str(row[13]),
            "raw_payload": json.dumps({
                "mes_lead": row[4],
                "tiempo_respuesta": row[12],
            }, default=str) if (row[4] is not None or row[12] is not None) else None,
        }
    wb.close()


def extract_lead_activities(filepath):
    """BaseLeads2 → raw.raw_lead_activities rows."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["BaseLeads2"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_num, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        # Fecha Actividad is a full datetime — split into date+time parts
        fecha_act = row[2]
        hora_act = None
        if isinstance(fecha_act, datetime):
            hora_act = str(fecha_act.hour)
        fecha_act_raw = _dt_str(fecha_act)

        # Fecha Planeada can be Excel serial number
        fecha_plan = row[9]
        hora_plan = None
        if isinstance(fecha_plan, (int, float)):
            dt = _excel_serial_to_datetime(fecha_plan)
            if dt:
                hora_plan = str(dt.hour)
                fecha_plan = dt
        fecha_plan_raw = _dt_str(fecha_plan)

        yield {
            "source_row_number": row_num,
            "aglead": _str(row[0]),
            "fecha_actividad_raw": fecha_act_raw,
            "hora_actividad_raw": hora_act,
            "fecha_programada_raw": fecha_plan_raw,
            "hora_programada_raw": hora_plan,
            "distribuidor": _str(row[1]),
            "grupo": _str(row[12]),
            "marca": _str(row[11]),
            "producto": _str(row[3]),
            "asesor": _str(row[5]),
            "campania": _str(row[7]),
            "subcampania": None,
            "fuente": _str(row[6]),
            "actividad": _str(row[8]),
            "estatus_actividad": _str(row[10]),
            "temperatura": _str(row[4]),
            "raw_payload": None,
        }
    wb.close()


def extract_finance_applications(filepath):
    """DBSolicitudes_Cetelem → raw.raw_finance_applications rows."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["DBSolicitudes_Cetelem"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_num, row in enumerate(rows, start=2):
        if not row or row[5] is None:  # folio is key
            continue
        yield {
            "source_row_number": row_num,
            "folio": _str(row[5]),
            "codigo_distribuidor": _str(row[18]),
            "distribuidor_ok": _str(row[4]),
            "nombre_distribuidor": _str(row[19]),
            "grupo": _str(row[20]),
            "marca": _str(row[21]),
            "oem": None,
            "producto": _str(row[24]),
            "carline": _str(row[12]),
            "modelo": None,
            "version": _str(row[13]),
            "anio_modelo_raw": _num_str(row[11]),
            "fecha_hora_recibida_raw": _str(row[7]),
            "fecha_ok_raw": _dt_str(row[0]),
            "periodo_yyyymm_raw": _num_str(row[6]),
            "tipo_vehiculo": _str(row[8]),
            "estatus": _str(row[9]),
            "subestatus": _str(row[10]),
            "tipo_persona": _str(row[22]),
            "plazo_meses_raw": _num_str(row[23]),
            "enganche_raw": _num_str(row[15]),
            "porcentaje_enganche_raw": _num_str(row[16]),
            "monto_unidad_raw": _num_str(row[14]),
            "monto_financiado_raw": _num_str(row[17]),
            "aprobado_raw": _num_str(row[1]),
            "base_flag_raw": _num_str(row[2]),
            "anf_flag_raw": _num_str(row[3]),
            "raw_payload": None,
        }
    wb.close()


def extract_market_prices(filepath):
    """DBPreciosMexico_ConMG → raw.raw_market_prices rows."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["DBPreciosMexico_ConMG"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_num, row in enumerate(rows, start=2):
        if not row or row[3] is None:  # modelo is key
            continue
        yield {
            "source_row_number": row_num,
            "mes_raw": _dt_str(row[0]),
            "marca": _str(row[2]),
            "oem": None,
            "modelo": _str(row[3]),
            "version": _str(row[4]),
            "anio_modelo_raw": _num_str(row[7]),
            "segmento_ryd": _str(row[1]),
            "tipo_carroceria": _str(row[5]),
            "puertas_raw": _num_str(row[6]),
            "precio_lista_raw": _num_str(row[8]),
            "precio_contado_neto_raw": _num_str(row[9]),
            "precio_financiado_raw": _num_str(row[10]),
            "moneda": _str(row[11]),
            "raw_payload": None,
        }
    wb.close()


def extract_sales(filepath):
    """DBVentas_ConMG → raw.raw_sales rows."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["DBVentas_ConMG"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_num, row in enumerate(rows, start=2):
        if not row or row[4] is None:
            continue
        yield {
            "source_row_number": row_num,
            "month_raw": _dt_str(row[0]),
            "dealer": _str(row[2]),
            "group_name": _str(row[1]),
            "brand": "MG",
            "oem": "MG MOTOR",
            "carline": _str(row[5]),
            "sale_type": _str(row[3]),
            "units_raw": _num_str(row[4]),
            "raw_payload": None,
        }
    wb.close()


def extract_claims(filepath):
    """DBSiniestros_Marsh → raw.raw_claims rows."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["DBSiniestros_Marsh"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_num, row in enumerate(rows, start=2):
        if not row or row[9] is None:  # numero_siniestro is key
            continue
        # Razón Social (row[5]) has no column in raw_claims — stash in raw_payload
        razon_social = row[5]
        yield {
            "source_row_number": row_num,
            "fecha_raw": _dt_str(row[0]),
            "poliza": _str(row[1]),
            "aseguradora": _str(row[2]),
            "programa": _str(row[3]),
            "nombre_agencia": _str(row[4]),
            "nombre_agencia_2": _str(row[6]),
            "grupo": _str(row[7]),
            "cliente": _str(row[8]),
            "numero_siniestro": _str(row[9]),
            "cobertura": _str(row[10]),
            "conductor": _str(row[11]),
            "telefono": _str(row[12]),
            "descripcion_vehiculo": _str(row[13]),
            "vin": _str(row[14]),
            "modelo": _str(row[15]),
            "ciudad": _str(row[16]),
            "estado": _str(row[17]),
            "raw_payload": json.dumps({"razon_social": razon_social}, default=str)
                if razon_social is not None else None,
        }
    wb.close()


def extract_inegi_sales(filepath):
    """BaseINEGIAutosLigerosMexico → raw.raw_inegi_sales rows."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["BaseINEGIAutosLigerosMexico"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row_num, row in enumerate(rows, start=2):
        if not row or row[3] is None:  # marca is key
            continue
        yield {
            "source_row_number": row_num,
            "tema": _str(row[0]),
            "anio_raw": _num_str(row[1]),
            "mes": _str(row[2]),
            "marca": _str(row[3]),
            "oem": None,
            "modelo": _str(row[4]),
            "tipo_vehiculo": _str(row[5]),
            "segmento": _str(row[6]),
            "tipo_origen": _str(row[7]),
            "pais_origen": _str(row[8]),
            "cantidad_raw": _num_str(row[9]),
            "raw_payload": None,
        }
    wb.close()


# Map of sheet name → (extractor function, raw table name).
# Sheets not listed here (Resumen, CatalogoTiendas ConMG) are skipped.
SHEET_MAP = {
    "BaseLeads1": (extract_leads, "raw.raw_leads"),
    "BaseLeads2": (extract_lead_activities, "raw.raw_lead_activities"),
    "DBSolicitudes_Cetelem": (extract_finance_applications, "raw.raw_finance_applications"),
    "DBPreciosMexico_ConMG": (extract_market_prices, "raw.raw_market_prices"),
    "DBVentas_ConMG": (extract_sales, "raw.raw_sales"),
    "DBSiniestros_Marsh": (extract_claims, "raw.raw_claims"),
    "BaseINEGIAutosLigerosMexico": (extract_inegi_sales, "raw.raw_inegi_sales"),
}
